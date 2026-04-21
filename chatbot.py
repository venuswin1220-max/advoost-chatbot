import streamlit as st
import pdfplumber
import openpyxl
import google.generativeai as genai
import os
from pathlib import Path

DOCS_DIR = Path(__file__).parent


def extract_pdf_text(pdf_path):
    text = f"\n\n=== 문서: {pdf_path.name} ===\n"
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text and page_text.strip():
                    text += f"\n[페이지 {i+1}]\n{page_text}\n"
    except Exception as e:
        text += f"(추출 오류: {e})\n"
    return text


def extract_excel_text(xlsx_path):
    text = f"\n\n=== 문서: {xlsx_path.name} ===\n"
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n[시트: {sheet_name}]\n"
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                row_str = " | ".join(cells)
                if row_str.replace("|", "").strip():
                    text += row_str + "\n"
        wb.close()
    except Exception as e:
        text += f"(추출 오류: {e})\n"
    return text


@st.cache_data(show_spinner=False)
def load_all_docs():
    all_text = ""
    for pdf in sorted(DOCS_DIR.glob("*.pdf")):
        all_text += extract_pdf_text(pdf)
    for xlsx in sorted(DOCS_DIR.glob("*.xlsx")):
        all_text += extract_excel_text(xlsx)
    return all_text


def get_api_key():
    try:
        return st.secrets["GEMINI_API_KEY"]
    except Exception:
        return os.environ.get("GEMINI_API_KEY", "")


def main():
    st.set_page_config(
        page_title="ADVoost Screen Q&A",
        page_icon="📺",
        layout="wide"
    )

    st.title("📺 ADVoost Screen 상품 Q&A")
    st.caption("소재 제작 가이드, 상품 스펙, FAQ에 대해 질문해 주세요.")

    with st.sidebar:
        st.header("⚙️ 설정")
        auto_key = get_api_key()
        if auto_key:
            api_key = auto_key
            st.success("API 키 연결됨")
        else:
            api_key = st.text_input(
                "Gemini API Key",
                type="password",
                placeholder="AIza...",
                help="aistudio.google.com에서 발급받은 API 키를 입력하세요"
            )

        if st.button("대화 초기화"):
            st.session_state.messages = []
            st.rerun()

        st.markdown("---")
        st.markdown("**참고 문서**")
        for f in sorted(DOCS_DIR.glob("*.pdf")):
            st.markdown(f"- {f.name}")
        for f in sorted(DOCS_DIR.glob("*.xlsx")):
            st.markdown(f"- {f.name}")

    if not api_key:
        st.info("왼쪽 사이드바에 **Gemini API Key**를 입력하면 시작됩니다.\n\n발급: [aistudio.google.com](https://aistudio.google.com) → Get API key")
        return

    with st.spinner("문서 로딩 중... (최초 1회)"):
        docs_content = load_all_docs()

    system_prompt = f"""당신은 네이버 ADVoost Screen 광고 상품 전문 안내 도우미입니다.
아웃소싱사 담당자가 소재 제작, 상품 스펙, FAQ 관련 질문을 할 때 정확하고 친절하게 답변해 주세요.

아래는 참고 문서 전체 내용입니다:
{docs_content}

답변 원칙:
- 문서 내용을 근거로 정확하게 답변합니다
- 해상도, 용량, 파일 포맷 등 수치는 반드시 정확하게 안내합니다
- 문서에 없는 내용은 "해당 내용은 제공된 문서에서 확인되지 않습니다"라고 안내합니다
- 한국어로 답변합니다
- CGV 영화 심의, 블랙화면 제거 등 특이사항은 반드시 강조해서 안내합니다"""

    if "messages" not in st.session_state:
        st.session_state.messages = []

    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.write(msg["content"])

    if user_input := st.chat_input("질문을 입력하세요 (예: CGV 소재 파일 용량 제한이 어떻게 되나요?)"):
        st.session_state.messages.append({"role": "user", "content": user_input})
        with st.chat_message("user"):
            st.write(user_input)

        with st.chat_message("assistant"):
            with st.spinner("답변 생성 중..."):
                try:
                    genai.configure(api_key=api_key)
                    model = genai.GenerativeModel(
                        model_name="gemini-1.5-flash-8b",
                        system_instruction=system_prompt
                    )
                    history = [
                        {
                            "role": "user" if m["role"] == "user" else "model",
                            "parts": [m["content"]]
                        }
                        for m in st.session_state.messages[:-1]
                    ]
                    chat = model.start_chat(history=history)
                    response = chat.send_message(user_input)
                    answer = response.text
                    st.write(answer)
                    st.session_state.messages.append({"role": "assistant", "content": answer})
                except Exception as e:
                    st.error(f"오류 발생: {e}")


if __name__ == "__main__":
    main()
